
import os
import imaplib
import email
from email.header import decode_header
import openpyxl
from datetime import datetime
import openai
import PyPDF2
from fuzzywuzzy import fuzz

# Email credentials (REPLACE WITH YOUR CREDENTIALS)
EMAIL_USER = "ingeniumprakash@zohomail.in"
EMAIL_PASSWORD = "pVLWzwcfTAkP"

# OpenAI API key
openai.api_key = "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"

# List of PDF file paths (REPLACE WITH YOUR PDF PATHS)
PDF_FILE_PATHS = ["./Adiabatic.pdf", "./Cooling.pdf", "./DSS.pdf"]

# Define sheet names and associated keywords
SHEET_KEYWORDS = {
    "adiabatic": ["adiabatic", "cooling", "system"],
    "ad": ["advertisement", "ad", "promotion"],
    "site visit": ["site visit", "meeting", "appointment"],
    "dust suppression": ["dust suppression", "dust control", "dust"],
    "fog cannon": ["fog cannon", "fogging", "fog", "cannon", "canon"],
    "Purchase Order": ["purchase order", "po", "order"],
    "Misting": ["misting", "mist"],
    "nozzle": ["nozzle", "spray nozzle", "nozzle tip"],
    "others": [],  # Default sheet for emails that don't match any keywords
    "unread": []  # For unread emails
}

# Canned response if the query does not relate to the PDF context
CANNED_RESPONSE = ("Thank you for reaching out. For further assistance, please contact our sales team at "
                   "917397180390 or visit our website https://truemist.in/")

# Define support-related keywords
SUPPORT_KEYWORDS = ["support", "help", "issue", "problem", "order status", "assistance", "query", "question"]

# Define phrases that indicate ChatGPT couldn't find the answer
UNANSWERED_PHRASES = [
    "I couldn't find",
    "doesn't seem to be mentioned",
    "not in the text provided",
    "no specific information",
    "not found in the text",
]


def connect_to_email():
    """Connects to the email server and logs in."""
    mail = imaplib.IMAP4_SSL("imap.zoho.in", 993)
    mail.login(EMAIL_USER, EMAIL_PASSWORD)
    print("✅ Zoho Mail login successful!")
    return mail


def fetch_emails(mail):
    """Fetches unread emails and marks them as read."""
    mail.select("inbox")
    _, messages = mail.search(None, "UNSEEN")
    email_ids = messages[0].split() if messages[0] else []

    if email_ids:
        for email_id in email_ids:
            mail.store(email_id, "+FLAGS", "\\Seen")  # Mark as read

    return email_ids


def get_email_details(mail, email_ids):
    """Extracts email details including attachments."""
    emails = []
    for email_id in email_ids:
        status, msg_data = mail.fetch(email_id, "(RFC822)")
        for response_part in msg_data:
            if isinstance(response_part, tuple):
                msg = email.message_from_bytes(response_part[1])
                subject, encoding = decode_header(msg["Subject"])[0]
                subject = subject.decode(encoding if encoding else "utf-8") if isinstance(subject, bytes) else subject
                sender = msg.get("From")
                date = msg.get("Date")
                body = ""
                has_attachment = "No"

                if msg.is_multipart():
                    for part in msg.walk():
                        content_type = part.get_content_type()
                        if content_type == "text/plain" and "attachment" not in part.get("Content-Disposition", ""):
                            body = part.get_payload(decode=True).decode(errors="ignore")
                        if part.get_filename():
                            has_attachment = "Yes"
                else:
                    body = msg.get_payload(decode=True).decode(errors="ignore")

                try:
                    timestamp = datetime.strptime(date, "%a, %d %b %Y %H:%M:%S %z").replace(tzinfo=None)
                except ValueError:
                    timestamp = datetime.now()

                emails.append((timestamp, sender, subject, body, has_attachment))
    return emails


def extract_text_from_pdfs(pdf_file_paths):
    """Extracts text from multiple PDF files and concatenates them."""
    combined_text = ""
    for pdf_file_path in pdf_file_paths:
        try:
            with open(pdf_file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                combined_text += "".join(page.extract_text() or "" for page in reader.pages) + "\n\n"
        except Exception as e:
            print(f"❌ Error extracting text from PDF {pdf_file_path}: {e}")
    return combined_text


def generate_chatgpt_response(email_body, pdf_text):
    """Generates a human-like response using ChatGPT with context from PDF text."""
    try:
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": (
                    "You are a friendly and knowledgeable assistant. "
                    "Please respond in a natural, human-like tone that is easy to understand with briefly, avoiding overly technical jargon."
                )},
                {"role": "user", "content": f"Context:\n{pdf_text[:10000]}\n\nQuery:\n{email_body}"},
            ],
            max_tokens=150,
            temperature=0.7,
        )
        return response.choices[0].message['content'].strip()
    except openai.error.OpenAIError:
        return "ChatGPT could not generate a response."


def find_existing_sender(wb, sender):
    """Searches all sheets for an existing sender and returns row + sheet name if found."""
    for sheet_name in SHEET_KEYWORDS.keys():
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            if ws[f"C{row}"].value == sender:
                return sheet_name, row, ws[f"A{row}"].value  # Return sheet name, row, and status
    return None, None, None  # If not found


def determine_target_sheet(body):
    """Determines the target sheet based on keywords in the email body using fuzzy matching."""
    body_lower = body.lower()

    best_match = ("others", 0)
    for sheet_name, keywords in SHEET_KEYWORDS.items():
        for keyword in keywords:
            # Calculate similarity score using fuzzy matching
            score = fuzz.partial_ratio(keyword.lower(), body_lower)
            if score > best_match[1]:
                best_match = (sheet_name, score)

    # Debug print to check the best match
    print(f"Best match: {best_match}")

    # Set a threshold to ensure relevance; if not reached, default to "others"
    return best_match[0] if best_match[1] >= 60 else "others"


def is_support_query(email_body):
    """Checks if the email body contains support-related keywords."""
    email_body_lower = email_body.lower()
    for keyword in SUPPORT_KEYWORDS:
        if keyword in email_body_lower:
            return True
    return False


def is_unanswered_query(chatgpt_response):
    """Checks if ChatGPT's response indicates that the specific information isn't found in the PDF."""
    chatgpt_response_lower = chatgpt_response.lower()
    for phrase in UNANSWERED_PHRASES:
        if phrase in chatgpt_response_lower:
            return True
    return False


def write_to_excel(data, filename="zoho_email_responses.xlsx"):
    """Writes email data to an Excel file and updates existing conversations."""
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.Workbook()
        for sheet in SHEET_KEYWORDS.keys():
            wb.create_sheet(title=sheet)
        del wb[wb.active.title]

    # Extract text from all PDFs once for context
    pdf_text = extract_text_from_pdfs(PDF_FILE_PATHS)

    for timestamp, sender, subject, body, has_attachment in data:
        formatted_time = timestamp.strftime("%d/%m/%Y %H:%M:%S")

        # Determine the target sheet based on the email body
        target_sheet = determine_target_sheet(body)

        # Check if the email is a support query
        if is_support_query(body):
            chatgpt_response = CANNED_RESPONSE  # Use the canned response for support queries
        elif target_sheet == "others":
            chatgpt_response = CANNED_RESPONSE  # Use the canned response if the query doesn't match the PDF context
        else:
            # Generate a ChatGPT response if the query matches the PDF context and is not a support query
            chatgpt_response = generate_chatgpt_response(body, pdf_text)
            # Check if ChatGPT's response indicates that the specific information isn't found in the PDF
            if is_unanswered_query(chatgpt_response):
                chatgpt_response = CANNED_RESPONSE  # Use the canned response

        # Check if sender exists in any sheet
        existing_sheet, existing_row, status = find_existing_sender(wb, sender)

        if existing_sheet and status == "Under Conversation":
            ws = wb[existing_sheet]

            # Update time and subject by appending new info
            ws.cell(row=existing_row, column=2).value = f"{formatted_time}, {ws.cell(row=existing_row, column=2).value}"
            ws.cell(row=existing_row, column=4).value = f"{subject}, {ws.cell(row=existing_row, column=4).value}"

            # Find next available query-response pair without gaps
            col = 6
            pair_count = 1
            while ws.cell(row=1, column=col).value is not None:
                col += 2
                pair_count += 1

            # Add new headers for the next query-response pair if needed
            ws.cell(row=1, column=col, value=f"Query {pair_count}")
            ws.cell(row=1, column=col + 1, value=f"Response {pair_count}")

            # Insert the new query and response
            ws.cell(row=existing_row, column=col, value=body)
            ws.cell(row=existing_row, column=col + 1, value=chatgpt_response)

        else:
            # Use the determined target sheet for new entries
            ws = wb[target_sheet]

            # If sheet is empty, create headers
            if ws.max_row == 1:
                ws.append(["Status", "Time", "Sender", "Subject", "Attachment", "Query 1", "Response 1"])

            # Insert new row at the top (Row 2 after headers)
            ws.insert_rows(2)
            ws["A2"] = "Under Conversation"
            ws["B2"] = formatted_time
            ws["C2"] = sender
            ws["D2"] = subject
            ws["E2"] = has_attachment
            ws["F2"] = body
            ws["G2"] = chatgpt_response

    wb.save(filename)
    wb.close()
    print(f"📂 Data successfully saved to {filename}")


def main():
    """Main function to fetch, process, and store emails in Excel efficiently."""
    mail = connect_to_email()
    email_ids = fetch_emails(mail)

    if not email_ids:
        print("⚠ No new unread emails to process.")
        return

    email_data = get_email_details(mail, email_ids)
    write_to_excel(email_data)

    mail.logout()


if __name__ == "__main__":
    main()